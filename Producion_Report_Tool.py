import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import traceback
import sys

class ProductionReporter:
    def __init__(self):
        self.root = tk.Tk()
        self.root.withdraw()

    def select_file(self):
        """Allow user to select input Excel file with error handling"""
        try:
            file_path = filedialog.askopenfilename(
                title="Select Production Data File",
                filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]
            )
            if not file_path:
                messagebox.showinfo("Info", "No file selected. Exiting.")
                sys.exit(0)
            return file_path
        except Exception as e:
            messagebox.showerror("File Selection Error", 
                                f"Error selecting file:\n{str(e)}")
            sys.exit(1)

    def safe_date_parse(self, date_str):
        """Handle WIP and other non-date values gracefully with better error handling"""
        try:
            if pd.isna(date_str) or str(date_str).strip().upper() in ['WIP', 'NA', 'N/A', '']:
                return pd.NaT
            # Try multiple date formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                try:
                    return pd.to_datetime(date_str, format=fmt)
                except:
                    continue
            return pd.to_datetime(date_str)  # Try pandas' automatic parsing
        except Exception as e:
            print(f"Warning: Could not parse date '{date_str}': {str(e)}")
            return pd.NaT

    def excel_column_to_index(self, col_letters):
        """Convert Excel column letters to zero-based index with validation"""
        try:
            index = 0
            for char in col_letters.upper():
                if not char.isalpha():
                    raise ValueError(f"Invalid column character: {char}")
                index = index * 26 + (ord(char) - ord('A')) + 1
            return index - 1
        except Exception as e:
            messagebox.showerror("Column Conversion Error", 
                                f"Error converting column '{col_letters}': {str(e)}")
            return None

    def validate_dataframe(self, df, required_cols, sheet_name=""):
        """Validate that dataframe contains required columns"""
        if df is None:
            return False
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            messagebox.showwarning("Missing Columns", 
                                  f"Sheet {sheet_name} is missing columns: {', '.join(missing_cols)}")
            return False
        return True

    def format_sheet(self, ws, headers, date_cols):
        """Apply consistent formatting to worksheets with error handling"""
        try:
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Format headers
            for col in ws.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
            
            # Format dates
            for col_name, df_col in date_cols.items():
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        if cell[0].value is not None:
                            cell[0].number_format = 'MM/DD/YYYY'
            
            # Auto-adjust column widths with limits
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        continue
                adjusted_width = min((max_length + 2) * 1.2, 50)  # Cap at 50
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
        except Exception as e:
            print(f"Warning: Error formatting sheet: {str(e)}")

    def add_production_chart(self, ws, df, report_type):
        """Add production volume charts to reports with error handling"""
        try:
            if len(df) < 1:
                return
                
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = f"{report_type} Production Volume"
            chart.y_axis.title = 'Records'
            chart.x_axis.title = report_type
            
            # Find record count columns (skip first column which is date)
            record_cols = [col for col in df.columns[1:] if 'Total Rec' in col]
            if not record_cols:
                return
            
            data_start_col = 2  # Skip first column (date)
            data_end_col = len(record_cols) + 1
            
            data = Reference(ws, min_col=data_start_col, max_col=data_end_col, 
                           min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, f"H{ws.max_row + 2}")
        except Exception as e:
            print(f"Warning: Error adding production chart: {str(e)}")

    def add_personnel_chart(self, ws, df):
        """Add horizontal bar chart for personnel performance with error handling"""
        try:
            if len(df) < 1:
                return
                
            chart = BarChart()
            chart.type = "bar"  # Horizontal bar chart
            chart.style = 10
            chart.title = "Top Performers by Total Records"
            chart.x_axis.title = 'Records'
            chart.y_axis.title = 'Personnel'
            
            # Get top 15 performers (or all if less than 15)
            top_count = min(15, len(df))
            
            # Data references
            data = Reference(ws, min_col=4, max_col=4, min_row=2, max_row=top_count+1)  # Total Rec. column
            cats = Reference(ws, min_col=2, min_row=2, max_row=top_count+1)  # Name column
            
            chart.add_data(data)
            chart.set_categories(cats)
            
            # Position chart to the right of the data
            ws.add_chart(chart, "F2")
        except Exception as e:
            print(f"Warning: Error adding personnel chart: {str(e)}")

    def add_personnel_trend_chart(self, ws, df, period_type):
        """Add line chart showing personnel performance trend with error handling"""
        try:
            if len(df) < 1:
                return
                
            top_names = df.groupby('Name')['Total Rec.'].sum().nlargest(5).index
            
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = f"Top Performers Trend ({period_type})"
            chart.y_axis.title = 'Records'
            chart.x_axis.title = period_type
            
            # Filter data for top performers
            filtered_df = df[df['Name'].isin(top_names)]
            if len(filtered_df) == 0:
                return
                
            pivot_data = filtered_df.pivot_table(
                index=period_type.lower(),
                columns='Name',
                values='Total Rec.',
                aggfunc='sum'
            ).fillna(0)
            
            # Add data to worksheet for charting
            start_row = ws.max_row + 1
            for r in dataframe_to_rows(pivot_data.reset_index(), index=False, header=True):
                ws.append(r)
            
            # Create chart
            data = Reference(ws, 
                           min_col=2, 
                           max_col=len(top_names)+1,
                           min_row=start_row,
                           max_row=ws.max_row)
            cats = Reference(ws,
                           min_col=1,
                           min_row=start_row + 1,
                           max_row=ws.max_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, f"H{ws.max_row + 2}")
        except Exception as e:
            print(f"Warning: Error adding trend chart: {str(e)}")

    def process_input_data(self, input_df):
        """Process all data including personnel performance reports with comprehensive error handling"""
        results = {
            'Key': {'df': None, 'cols': ['J', 'K', 'L', 'M', 'N'], 
                   'names': ['Key Branch', 'Outdate', 'Duedate', 'Indate', 'Total Rec.']},
            'QC': {'df': None, 'cols': ['U', 'V', 'W', 'X'], 
                  'names': ['QC Branch', 'Outdate', 'Indate', 'Total Rec.']},
            'Final': {'df': None, 'cols': ['X', 'Y', 'Z', 'AA', 'AB', 'AC'],
                     'names': ['Total Rec.', 'Final Person', 'Outdate', 'Indate', 'Status', 'Shipment Date']},
            'Weekly': {'df': None},
            'Monthly': {'df': None},
            'Personnel': {'df': None},
            'PersonnelWeekly': {'df': None},
            'PersonnelMonthly': {'df': None}
        }
        
        try:
            # Validate input dataframe
            if input_df.empty:
                messagebox.showerror("Input Error", "The selected file is empty.")
                return None
            

            # Process Key data (columns J-N)
            try:
                key_cols = [self.excel_column_to_index(col) for col in results['Key']['cols']]
                if None in key_cols:
                    raise ValueError("Invalid column specification for Key data")
                    
                key_df = input_df.iloc[:, key_cols].copy()
                key_df.columns = results['Key']['names']
                
                for date_col in ['Outdate', 'Duedate', 'Indate']:
                    key_df[date_col] = key_df[date_col].apply(self.safe_date_parse)
                
                key_df = key_df.dropna(subset=['Key Branch', 'Total Rec.'])
                if not key_df.empty:
                    key_df['Processing Days'] = (key_df['Indate'] - key_df['Outdate']).dt.days
                    key_df['On Time Status'] = (key_df['Indate'] <= key_df['Duedate']).map({True: 'On Time', False: 'Delayed'})
                    key_df['Week'] = key_df['Indate'].dt.to_period('W').dt.start_time
                    key_df['Month'] = key_df['Indate'].dt.to_period('M').dt.start_time
                    key_df['Process'] = 'Key'
                
                results['Key']['df'] = key_df
            except Exception as e:
                messagebox.showwarning("Key Data Warning", 
                                      f"Error processing Key data: {str(e)}\nContinuing without Key data.")
                results['Key']['df'] = pd.DataFrame()

            # Process QC data (columns U-X)
            try:
                qc_cols = [self.excel_column_to_index(col) for col in results['QC']['cols']]
                if None in qc_cols:
                    raise ValueError("Invalid column specification for QC data")
                    
                qc_df = input_df.iloc[:, qc_cols].copy()
                qc_df.columns = results['QC']['names']
                
                for date_col in ['Outdate', 'Indate']:
                    qc_df[date_col] = qc_df[date_col].apply(self.safe_date_parse)
                
                qc_df = qc_df.dropna(subset=['QC Branch', 'Total Rec.'])
                if not qc_df.empty:
                    qc_df['Processing Days'] = (qc_df['Indate'] - qc_df['Outdate']).dt.days
                    qc_df['Week'] = qc_df['Indate'].dt.to_period('W').dt.start_time
                    qc_df['Month'] = qc_df['Indate'].dt.to_period('M').dt.start_time
                    qc_df['Process'] = 'QC'
                
                results['QC']['df'] = qc_df
            except Exception as e:
                messagebox.showwarning("QC Data Warning", 
                                      f"Error processing QC data: {str(e)}\nContinuing without QC data.")
                results['QC']['df'] = pd.DataFrame()

            # Process Final data (columns X-AC)
            try:
                final_cols = [self.excel_column_to_index(col) for col in results['Final']['cols']]
                if None in final_cols:
                    raise ValueError("Invalid column specification for Final data")
                    
                final_df = input_df.iloc[:, final_cols].copy()
                final_df.columns = results['Final']['names']
                
                for date_col in ['Outdate', 'Indate', 'Shipment Date']:
                    final_df[date_col] = final_df[date_col].apply(self.safe_date_parse)
                
                final_df = final_df.dropna(subset=['Final Person', 'Total Rec.'])
                if not final_df.empty:
                    final_df['Processing Days'] = (final_df['Shipment Date'] - final_df['Outdate']).dt.days
                    final_df['Week'] = final_df['Shipment Date'].dt.to_period('W').dt.start_time
                    final_df['Month'] = final_df['Shipment Date'].dt.to_period('M').dt.start_time
                    final_df['Process'] = 'Final'
                
                results['Final']['df'] = final_df
            except Exception as e:
                messagebox.showwarning("Final Data Warning", 
                                     f"Error processing Final data: {str(e)}\nContinuing without Final data.")
                results['Final']['df'] = pd.DataFrame()

            # Generate personnel data
            personnel_data = []
            personnel_weekly = []
            personnel_monthly = []
            
            # Process Key personnel
            if results['Key']['df'] is not None and not results['Key']['df'].empty:
                key_personnel = results['Key']['df'].copy()
                personnel_data.append(key_personnel[['Key Branch', 'Total Rec.', 'Process']].rename(columns={'Key Branch': 'Name'}))
                
                # Weekly
                key_weekly = key_personnel.groupby(['Week', 'Key Branch']).agg({
                    'Total Rec.': 'sum'
                }).reset_index()
                key_weekly['Process'] = 'Key'
                key_weekly = key_weekly.rename(columns={'Key Branch': 'Name'})
                personnel_weekly.append(key_weekly)
                
                # Monthly
                key_monthly = key_personnel.groupby(['Month', 'Key Branch']).agg({
                    'Total Rec.': 'sum'
                }).reset_index()
                key_monthly['Process'] = 'Key'
                key_monthly = key_monthly.rename(columns={'Key Branch': 'Name'})
                personnel_monthly.append(key_monthly)
            
            # Process QC personnel
            if results['QC']['df'] is not None and not results['QC']['df'].empty:
                qc_personnel = results['QC']['df'].copy()
                personnel_data.append(qc_personnel[['QC Branch', 'Total Rec.', 'Process']].rename(columns={'QC Branch': 'Name'}))
                
                # Weekly
                qc_weekly = qc_personnel.groupby(['Week', 'QC Branch']).agg({
                    'Total Rec.': 'sum'
                }).reset_index()
                qc_weekly['Process'] = 'QC'
                qc_weekly = qc_weekly.rename(columns={'QC Branch': 'Name'})
                personnel_weekly.append(qc_weekly)
                
                # Monthly
                qc_monthly = qc_personnel.groupby(['Month', 'QC Branch']).agg({
                    'Total Rec.': 'sum'
                }).reset_index()
                qc_monthly['Process'] = 'QC'
                qc_monthly = qc_monthly.rename(columns={'QC Branch': 'Name'})
                personnel_monthly.append(qc_monthly)
            
            # Process Final personnel
            if results['Final']['df'] is not None and not results['Final']['df'].empty:
                final_personnel = results['Final']['df'].copy()
                personnel_data.append(final_personnel[['Final Person', 'Total Rec.', 'Process']].rename(columns={'Final Person': 'Name'}))
                
                # Weekly
                final_weekly = final_personnel.groupby(['Week', 'Final Person']).agg({
                    'Total Rec.': 'sum'
                }).reset_index()
                final_weekly['Process'] = 'Final'
                final_weekly = final_weekly.rename(columns={'Final Person': 'Name'})
                personnel_weekly.append(final_weekly)
                
                # Monthly
                final_monthly = final_personnel.groupby(['Month', 'Final Person']).agg({
                    'Total Rec.': 'sum'
                }).reset_index()
                final_monthly['Process'] = 'Final'
                final_monthly = final_monthly.rename(columns={'Final Person': 'Name'})
                personnel_monthly.append(final_monthly)
            
            # Combine all personnel data
            if personnel_data:
                try:
                    personnel_report = pd.concat(personnel_data)
                    personnel_report = personnel_report.groupby(['Name', 'Process'])['Total Rec.'].sum().reset_index()
                    personnel_report = personnel_report.sort_values(by='Total Rec.', ascending=False)
                    personnel_report.insert(0, 'S.No.', range(1, len(personnel_report) + 1))
                    results['Personnel']['df'] = personnel_report
                except Exception as e:
                    messagebox.showwarning("Personnel Report Warning", 
                                         f"Error creating personnel report: {str(e)}")
                    results['Personnel']['df'] = pd.DataFrame()
            
            # Combine weekly personnel data
            if personnel_weekly:
                try:
                    personnel_weekly_report = pd.concat(personnel_weekly)
                    results['PersonnelWeekly']['df'] = personnel_weekly_report.sort_values(['Week', 'Total Rec.'], ascending=[True, False])
                except Exception as e:
                    messagebox.showwarning("Weekly Personnel Warning", 
                                         f"Error creating weekly personnel report: {str(e)}")
                    results['PersonnelWeekly']['df'] = pd.DataFrame()
            
            # Combine monthly personnel data
            if personnel_monthly:
                try:
                    personnel_monthly_report = pd.concat(personnel_monthly)
                    results['PersonnelMonthly']['df'] = personnel_monthly_report.sort_values(['Month', 'Total Rec.'], ascending=[True, False])
                except Exception as e:
                    messagebox.showwarning("Monthly Personnel Warning", 
                                         f"Error creating monthly personnel report: {str(e)}")
                    results['PersonnelMonthly']['df'] = pd.DataFrame()

            # Generate process weekly reports
            weekly_data = []
            for process in ['Key', 'QC', 'Final']:
                df = results[process]['df']
                if df is not None and not df.empty:
                    try:
                        weekly = df.groupby(['Week', 'Process']).agg({
                            'Total Rec.': 'sum',
                            'Processing Days': 'mean'
                        }).reset_index()
                        weekly_data.append(weekly)
                    except Exception as e:
                        print(f"Warning: Error aggregating weekly {process} data: {str(e)}")
            
            if weekly_data:
                try:
                    weekly_report = pd.concat(weekly_data)
                    weekly_report = weekly_report.pivot_table(
                        index='Week',
                        columns='Process',
                        values=['Total Rec.', 'Processing Days'],
                        aggfunc='sum'
                    ).fillna(0)
                    weekly_report.columns = [' '.join(col).strip() for col in weekly_report.columns.values]
                    results['Weekly']['df'] = weekly_report.reset_index()
                except Exception as e:
                    messagebox.showwarning("Weekly Report Warning", 
                                         f"Error creating weekly report: {str(e)}")
                    results['Weekly']['df'] = pd.DataFrame()
            
            # Generate monthly report
            monthly_data = []
            for process in ['Key', 'QC', 'Final']:
                df = results[process]['df']
                if df is not None and not df.empty:
                    try:
                        monthly = df.groupby(['Month', 'Process']).agg({
                            'Total Rec.': 'sum',
                            'Processing Days': 'mean'
                        }).reset_index()
                        monthly_data.append(monthly)
                    except Exception as e:
                        print(f"Warning: Error aggregating monthly {process} data: {str(e)}")
            
            if monthly_data:
                try:
                    monthly_report = pd.concat(monthly_data)
                    monthly_report = monthly_report.pivot_table(
                        index='Month',
                        columns='Process',
                        values=['Total Rec.', 'Processing Days'],
                        aggfunc='sum'
                    ).fillna(0)
                    monthly_report.columns = [' '.join(col).strip() for col in monthly_report.columns.values]
                    results['Monthly']['df'] = monthly_report.reset_index()
                except Exception as e:
                    messagebox.showwarning("Monthly Report Warning", 
                                         f"Error creating monthly report: {str(e)}")
                    results['Monthly']['df'] = pd.DataFrame()
            
            return results
        
        except Exception as e:
            messagebox.showerror("Processing Error", 
                               f"Critical error processing data:\n{str(e)}\n{traceback.format_exc()}")
            return None

    def save_reports(self, processed_data):
        """Save all reports to Excel with comprehensive error handling"""
        if not processed_data:
            messagebox.showerror("Save Error", "No processed data to save.")
            return
        
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Helper function to safely add data to worksheet
            def safe_add_data(ws, data, headers=None):
                try:
                    if headers:
                        ws.append(headers)
                    for row in data:
                        ws.append(row)
                    return True
                except Exception as e:
                    print(f"Warning: Error adding data to sheet {ws.title}: {str(e)}")
                    return False
            
            # Create process sheets
            for process_name in ['Key', 'QC', 'Final']:
                data = processed_data[process_name]
                if data['df'] is not None and not data['df'].empty:
                    try:
                        ws = wb.create_sheet(title=f"{process_name} Process")
                        
                        if process_name == 'Key':
                            headers = ['S.No.', 'Key Branch', 'Start Date', 'Due Date', 
                                     'End Date', 'Total Records', 'Processing Days', 'Status']
                            date_cols = {'Start Date': 'Outdate', 'Due Date': 'Duedate', 'End Date': 'Indate'}
                            rows = []
                            for i, (_, row) in enumerate(data['df'].iterrows(), 1):
                                rows.append([
                                    i, row['Key Branch'], row['Outdate'], row['Duedate'],
                                    row['Indate'], row['Total Rec.'], row['Processing Days'], row['On Time Status']
                                ])
                        elif process_name == 'QC':
                            headers = ['S.No.', 'QC Branch', 'Start Date', 'End Date', 
                                     'Total Records', 'Processing Days']
                            date_cols = {'Start Date': 'Outdate', 'End Date': 'Indate'}
                            rows = []
                            for i, (_, row) in enumerate(data['df'].iterrows(), 1):
                                rows.append([
                                    i, row['QC Branch'], row['Outdate'], row['Indate'],
                                    row['Total Rec.'], row['Processing Days']
                                ])
                        else:
                            headers = ['S.No.', 'Final Person', 'Start Date', 'QC End Date',
                                     'Shipment Date', 'Total Records', 'Processing Days', 'Status']
                            date_cols = {'Start Date': 'Outdate', 'QC End Date': 'Indate', 'Shipment Date': 'Shipment Date'}
                            rows = []
                            for i, (_, row) in enumerate(data['df'].iterrows(), 1):
                                rows.append([
                                    i, row['Final Person'], row['Outdate'], row['Indate'],
                                    row['Shipment Date'], row['Total Rec.'], row['Processing Days'], row['Status']
                                ])
                        
                        if safe_add_data(ws, rows, headers):
                            self.format_sheet(ws, headers, date_cols)
                    except Exception as e:
                        messagebox.showwarning(f"{process_name} Sheet Warning", 
                                            f"Error creating {process_name} sheet: {str(e)}")
            
            # Create Weekly/Monthly reports
            for report_type in ['Weekly', 'Monthly']:
                data = processed_data[report_type]
                if data['df'] is not None and not data['df'].empty:
                    try:
                        ws = wb.create_sheet(title=f"{report_type} Report")
                        rows = []
                        for r in dataframe_to_rows(data['df'], index=False, header=True):
                            rows.append(r)
                        
                        if safe_add_data(ws, rows):
                            self.format_sheet(ws, list(data['df'].columns), {})
                            self.add_production_chart(ws, data['df'], report_type)
                    except Exception as e:
                        messagebox.showwarning(f"{report_type} Report Warning", 
                                              f"Error creating {report_type} report: {str(e)}")
            
            # Create Personnel reports
            if processed_data['Personnel']['df'] is not None and not processed_data['Personnel']['df'].empty:
                try:
                    ws_personnel = wb.create_sheet(title="Personnel Performance")
                    headers = ['S.No.', 'Name', 'Process', 'Total Records']
                    rows = []
                    for _, row in processed_data['Personnel']['df'].iterrows():
                        rows.append([row['S.No.'], row['Name'], row['Process'], row['Total Rec.']])
                    
                    if safe_add_data(ws_personnel, rows, headers):
                        self.format_sheet(ws_personnel, headers, {})
                        self.add_personnel_chart(ws_personnel, processed_data['Personnel']['df'])
                except Exception as e:
                    messagebox.showwarning("Personnel Sheet Warning", 
                                         f"Error creating personnel sheet: {str(e)}")
            
            if processed_data['PersonnelWeekly']['df'] is not None and not processed_data['PersonnelWeekly']['df'].empty:
                try:
                    ws_personnel_weekly = wb.create_sheet(title="Personnel Weekly")
                    headers = ['Week', 'Name', 'Process', 'Total Records']
                    rows = []
                    for _, row in processed_data['PersonnelWeekly']['df'].iterrows():
                        rows.append([
                            row['Week'].strftime('%Y-%m-%d'),
                            row['Name'],
                            row['Process'],
                            row['Total Rec.']
                        ])
                    
                    if safe_add_data(ws_personnel_weekly, rows, headers):
                        self.format_sheet(ws_personnel_weekly, headers, {'Week': 'Week'})
                        self.add_personnel_trend_chart(ws_personnel_weekly, processed_data['PersonnelWeekly']['df'], "Weekly")
                except Exception as e:
                    messagebox.showwarning("Weekly Personnel Sheet Warning", 
                                         f"Error creating weekly personnel sheet: {str(e)}")
            
            if processed_data['PersonnelMonthly']['df'] is not None and not processed_data['PersonnelMonthly']['df'].empty:
                try:
                    ws_personnel_monthly = wb.create_sheet(title="Personnel Monthly")
                    headers = ['Month', 'Name', 'Process', 'Total Records']
                    rows = []
                    for _, row in processed_data['PersonnelMonthly']['df'].iterrows():
                        rows.append([
                            row['Month'].strftime('%Y-%m'),
                            row['Name'],
                            row['Process'],
                            row['Total Rec.']
                        ])
                    
                    if safe_add_data(ws_personnel_monthly, rows, headers):
                        self.format_sheet(ws_personnel_monthly, headers, {'Month': 'Month'})
                        self.add_personnel_trend_chart(ws_personnel_monthly, processed_data['PersonnelMonthly']['df'], "Monthly")
                except Exception as e:
                    messagebox.showwarning("Monthly Personnel Sheet Warning", 
                                        f"Error creating monthly personnel sheet: {str(e)}")
            
            # Save file
            output_file = "Production_Performance_Report.xlsx"
            try:
                wb.save(output_file)
                messagebox.showinfo("Success", f"Report successfully saved as {output_file}")
            except PermissionError:
                messagebox.showerror("Save Error", 
                                   f"Could not save {output_file}. The file may be open in another program.")
            except Exception as e:
                messagebox.showerror("Save Error", 
                                  f"Error saving file: {str(e)}")
        
        except Exception as e:
            messagebox.showerror("Report Generation Error", 
                              f"Critical error generating reports:\n{str(e)}")

    def run(self):
        """Main execution method with full error handling"""
        try:
            input_file = self.select_file()
            if not input_file:
                return
            
            try:
                input_df = pd.read_excel(input_file)
            except Exception as e:
                messagebox.showerror("File Read Error", 
                                    f"Error reading input file:\n{str(e)}")
                return
            
            processed_data = self.process_input_data(input_df)
            
            if processed_data is not None:
                self.save_reports(processed_data)
            
        except Exception as e:
            messagebox.showerror("Application Error", 
                               f"Unexpected error:\n{str(e)}\n{traceback.format_exc()}")
        finally:
            self.root.quit()

if __name__ == "__main__":
    app = ProductionReporter()
    app.run()
